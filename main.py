import os
from tkinter import *
from tkinter import filedialog
from tkinter.ttk import Combobox

from openpyxl import *
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

path = "C:\\Users\\snoew\\OneDrive\\Skrivbord\\Projekt\\Accounting_Test.xlsx"
wb = load_workbook(path)


def sheets():
    months = ["Januari", "Februari", "Mars", "April", "Maj", "Juni",
              "Juli", "Augusti", "September", "Oktober", "November", "December"]
    headers = ['Dag', 'Köpare, Säljare, Varuslag, etc.', 'Verif.\nnr', 'Inbetalningar', 'Utbetalningar',
               'Utgående\nmoms',
               'Inventarier\nförsäljning', 'Nötkreatur', 'Svin', 'Skog och\nskogs-\nprodukter',
               'Övriga \ninbetalningar',
               'Ingående\nmoms', 'Inventarier\ninköp', 'Inköp djur', 'Omkostnader\nskogen', 'Omkostnader\ndjurskötseln',
               'Drivmedel,\neldnings och\nsmörolja', 'Underhåll\ninventarier',
               'Kontors-\nkostnader,\nbokföring,\ntelefon',
               'Försäkrings\npremier', 'Övriga\nutbetalningar',
               ' Underhåll\nnärings-\nfastigheter\nekonomi-\nbyggnader',
               'Underhåll\nnärings-\nfastigheter\nbostäder\n(inkl moms)', 'Underhåll\nmark-\nanläggning']

    for m in range(len(months)):
        temp = months[m]
        sheets = wb.create_sheet(title=temp)

        # Set Dimensions
        sheets.column_dimensions['A'].width = 5
        sheets.column_dimensions['B'].width = 30
        sheets.column_dimensions['C'].width = 7
        column = 4
        while column < 25:
            i = get_column_letter(column)
            sheets.column_dimensions[i].width = 15
            column += 1

        # Merge cells
        sheets.merge_cells('A1:F1')
        sheets.merge_cells('A3:B3')
        sheets.merge_cells('G1:K1')
        sheets.merge_cells('L1:X1')

        # Give the headers values
        sheets['A1'].value = "Fördelning av"
        sheets['A3'].value = "SUMMA"
        sheets['C3'].value = "---"
        sheets['G1'].value = "Inbetalningar"
        sheets['L1'].value = "Utbetalningar"
        for h in range(len(headers)):
            tempHead = headers[h]
            sheets.cell(row=2, column=h + 1).value = tempHead

        # Formula for the total
        sheets.cell(row=3, column=4).value = '=SUM(D4:INDEX(D:D,ROWS(D:D)))'
        sheets.cell(row=3, column=5).value = '=SUM(E4:INDEX(E:E,ROWS(E:E)))'
        sheets.cell(row=3, column=6).value = '=SUM(F4:INDEX(F:F,ROWS(F:F)))'
        sheets.cell(row=3, column=7).value = '=SUM(G4:INDEX(G:G,ROWS(G:G)))'
        sheets.cell(row=3, column=8).value = '=SUM(H4:INDEX(H:H,ROWS(H:H)))'
        sheets.cell(row=3, column=9).value = '=SUM(I4:INDEX(I:I,ROWS(I:I)))'
        sheets.cell(row=3, column=10).value = '=SUM(J4:INDEX(J:J,ROWS(J:J)))'
        sheets.cell(row=3, column=11).value = '=SUM(K4:INDEX(K:K,ROWS(K:K)))'
        sheets.cell(row=3, column=12).value = '=SUM(L4:INDEX(L:L,ROWS(L:L)))'
        sheets.cell(row=3, column=13).value = '=SUM(M4:INDEX(M:M,ROWS(M:M)))'
        sheets.cell(row=3, column=14).value = '=SUM(N4:INDEX(N:N,ROWS(N:N)))'
        sheets.cell(row=3, column=15).value = '=SUM(O4:INDEX(O:O,ROWS(O:O)))'
        sheets.cell(row=3, column=16).value = '=SUM(P4:INDEX(P:P,ROWS(P:P)))'
        sheets.cell(row=3, column=17).value = '=SUM(Q4:INDEX(Q:Q,ROWS(Q:Q)))'
        sheets.cell(row=3, column=18).value = '=SUM(R4:INDEX(R:R,ROWS(R:R)))'
        sheets.cell(row=3, column=19).value = '=SUM(S4:INDEX(S:S,ROWS(S:S)))'
        sheets.cell(row=3, column=20).value = '=SUM(T4:INDEX(T:T,ROWS(T:T)))'
        sheets.cell(row=3, column=21).value = '=SUM(U4:INDEX(U:U,ROWS(U:U)))'
        sheets.cell(row=3, column=22).value = '=SUM(V4:INDEX(V:V,ROWS(V:V)))'
        sheets.cell(row=3, column=23).value = '=SUM(W4:INDEX(W:W,ROWS(W:W)))'
        sheets.cell(row=3, column=24).value = '=SUM(W4:INDEX(W:W,ROWS(W:W)))'


        # Center the values
        wb.alignment = Alignment(horizontal='left')
        for cell in sheets[1]:
            cell.alignment = Alignment(horizontal='center')
        for cell in sheets[2]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        sheets['C3'].alignment = Alignment(horizontal='center')

    del wb["Sheet1"]


def excel():
    thin = Side(border_style="thin", color="000000")
    for row in wb.active:
        for cell in row:
            cell.border = Border(left=thin, right=thin)

    # Background color
    current_row = wb.active.max_row
    if current_row % 2 == 0:
        for cell in wb.active[current_row:current_row]:
            cell.fill = PatternFill(start_color="ADD8E6", fill_type="solid")


    # Set the font
    wb.font = Font(size=12)
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    for cell in wb.active['K']:
        cell.border = Border(bottom=thin, top=thin, left=thin, right=thick)
    for cell in wb.active['C']:
        cell.border = Border(bottom=thin, top=thin, left=thin, right=thick)
    for row in wb.active:
        for cell in row:
            cell.border = Border(left=thin, right=thin)
    for cell in wb.active[1:1]:
        cell.font = Font(size=14, bold=True)
    for cell in wb.active[2:2]:
        cell.border = Border(bottom=thin, left=thin, right=thin, top=thin)
    for cell in wb.active[3:3]:
        cell.font = Font(size=14, bold=True)
        cell.border = Border(bottom=thick, top=thin, left=thin, right=thin)
    wb.active.cell(row=1, column=11).border = Border(bottom=thin, top=thin, left=thin, right=thick)
    wb.active.cell(row=2, column=11).border = Border(bottom=thin, top=thin, left=thin, right=thick)
    wb.active.cell(row=3, column=11).border = Border(bottom=thick, top=thin, left=thin, right=thick)



# Set focus(event) for every field


# Save file
# def saveFile():
# if path == None:
# path = asksaveasfilename(initialfile='Untitled.xlsx', defaultextension=".xlsx",
# filetypes=[("All Files", "*.*"),("Excel Documents","*.xlsx")])
# if path == "":
# path = None
# else:
## Try to save the file
# file = open(path,"w")
# file.write(

# def showFile()

# Driver code
if __name__ == "__main__":
    def set_jan():
        month = months_headers[0]
        wb.active = 0
        receipt_entries(month)


    def set_feb():
        month = months_headers[1]
        wb.active = 1
        receipt_entries(month)


    def set_mar():
        month = months_headers[2]
        wb.active = 2
        receipt_entries(month)


    def set_apr():
        month = months_headers[3]
        wb.active = 3
        receipt_entries(month)


    def set_may():
        month = months_headers[4]
        wb.active = 4
        receipt_entries(month)


    def set_jun():
        month = months_headers[5]
        wb.active = 5
        receipt_entries(month)


    def set_jul():
        month = months_headers[6]
        wb.active = 6
        receipt_entries(month)


    def set_aug():
        month = months_headers[7]
        wb.active = 7
        receipt_entries(month)


    def set_sep():
        month = months_headers[8]
        wb.active = 8
        receipt_entries(month)


    def set_oct():
        month = months_headers[9]
        wb.active = 9
        receipt_entries(month)


    def set_nov():
        month = months_headers[10]
        wb.active = 10
        receipt_entries(month)


    def set_dec():
        month = months_headers[11]
        wb.active = 11
        receipt_entries(month)


    def browseFiles():
        filename = filedialog.askopenfilename(initialdir="/", title="Select a file",
                                              filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
        wb = load_workbook(filename)
        root.title(os.path.basename(filename))


    def newFileWindow():
        """def newfile():
            wb = Workbook()
            filepath = f"C:\\Users\\snoew\\OneDrive\\Skrivbord\\Projekt\\{entry}.xlxs"
            wb.save(filepath)
            #sheets()
            nf.destroy()"""
        nf = Toplevel(root)
        nf.geometry("300x300")
        nf.configure(bg='light blue')
        Default_Label(nf, text="Not yet implemented").pack()
        Default_Button(nf, text="Okej", command=nf.destroy).pack()
        """Default_Label(nf, text="Välj namn på filen: ").pack()
        entry = Entry(nf, font=12).pack()
        Default_Button(nf, text="Okej!", command=newfile).pack()"""

    def clearFrameMonths():
        for widget in mainframe.winfo_children():
            widget.pack_forget()

    def clearFrameEntries():
        for frame in mainframe.winfo_children():
            frame.pack_forget()
            for widget in mainframe.winfo_children():
                widget.pack_forget()
            for widget in frame2_right.winfo_children():
                widget.pack_forget()
            for widget in frame2_left.winfo_children():
                widget.pack_forget()
            for widget in frame3_left_bot.winfo_children():
                widget.pack_forget()
            for widget in frame3_left_top.winfo_children():
                widget.pack_forget()
            for widget in frame4_bot.winfo_children():
                widget.pack_forget()
            for widget in frame4_top.winfo_children():
                widget.pack_forget()




    root = Tk()
    menu = Menu(root)
    root.config(menu=menu)
    filemenu = Menu(menu)
    menu.add_cascade(label='File', menu=filemenu)
    filemenu.add_command(label='New', command=newFileWindow)
    filemenu.add_command(label='Open...', command=browseFiles)
    filemenu.add_separator()
    filemenu.add_command(label='Exit', command=root.quit)

    # Set title of GUI Window
    root.title(os.path.basename(path))
    # Set the config of GUI window
    root.geometry("500x550")
    root.configure(bg='light blue')

    sub_type_headers1 = ['----', 'Inventarier försäljning', 'Nötkreatur', 'Svin', 'Skog och skogsprodukter',
                         'Övriga inbetalningar']
    sub_type_headers2 = ['----', 'Inventarier inköp', 'Inköp djur', 'Omkostnader skogen', 'Omkostnader djurskötseln',
                         'Drivmedel, eldnings och smörolja', 'Underhåll inventarier',
                         'Kontorskostnader, bokföring, telefon',
                         'Försäkrings premier', 'Övriga utbetalningar',
                         ' Underhåll näringsfastigheter ekonomibyggnader',
                         'Underhåll näringsfastigheter bostäder (inkl moms)', 'Underhåll markanläggning']
    months_headers = ["Januari", "Februari", "Mars", "April", "Maj", "Juni",
                      "Juli", "Augusti", "September", "Oktober", "November", "December"]

    # Create frame
    mainframe = Frame(root, bg='light blue')
    mainframe.pack(fill="both", side=TOP)

    frame2 = Frame(root, bg='light blue')
    frame2.pack(fill="both", side=TOP)
    frame2_left = Frame(frame2, bg='light blue')
    frame2_left.pack(side=LEFT, fill="both")
    frame2_right = Frame(frame2, bg='light blue')
    frame2_right.pack(side=RIGHT, fill="both")

    frame3 = Frame(root, bg='light blue')
    frame3.pack(fill="both", side=TOP)
    frame3_left_top = Frame(frame3, bg='light blue')
    frame3_left_top.pack(side=TOP, fill="both")
    frame3_left_bot = Frame(frame3, bg='light blue')
    frame3_left_bot.pack(side=TOP, fill="both")

    frame4 = Frame(root, bg='light blue')
    frame4.pack(expand=True, fill="both", side=TOP)
    frame4_top = Frame(frame4, bg='light blue')
    frame4_top.pack(side=TOP, fill="both")
    frame4_bot = Frame(frame4, bg='light blue')
    frame4_bot.pack(side=TOP, fill="both")


    class Default_Label(Label):
        def __init__(self, *args, **kwargs):
            Label.__init__(self, *args, **kwargs)
            self['bg'] = 'light blue'
            self['font'] = 12


    class Default_Button(Button):
        def __init__(self, *args, **kwargs):
            Button.__init__(self, *args, **kwargs)
            self['font'] = 14
            self['width'] = 15

    def test():
        clearFrameEntries()
        months()
    def months():
        clearFrameEntries()
        Default_Label(mainframe, text='Välj månad').pack(side="top", pady=5)
        Default_Button(mainframe, text=months_headers[0], command=set_jan).pack(pady=5)
        Default_Button(mainframe, text=months_headers[1], command=set_feb).pack(pady=5)
        Default_Button(mainframe, text=months_headers[2], command=set_mar).pack(pady=5)
        Default_Button(mainframe, text=months_headers[3], command=set_apr).pack(pady=5)
        Default_Button(mainframe, text=months_headers[4], command=set_may).pack(pady=5)
        Default_Button(mainframe, text=months_headers[5], command=set_jun).pack(pady=5)
        Default_Button(mainframe, text=months_headers[6], command=set_jul).pack(pady=5)
        Default_Button(mainframe, text=months_headers[7], command=set_aug).pack(pady=5)
        Default_Button(mainframe, text=months_headers[8], command=set_sep).pack(pady=5)
        Default_Button(mainframe, text=months_headers[9], command=set_oct).pack(pady=5)
        Default_Button(mainframe, text=months_headers[10], command=set_nov).pack(pady=5)
        Default_Button(mainframe, text=months_headers[11], command=set_dec).pack(pady=5)








    def receipt_entries(month):
        clearFrameMonths()

        def focus1(event):
            brutto_field.focus_set()

        def focus2(event):
            momsPercent_field.focus_set()

        def clear():
            # Clear every field in the GUI
            receipt_name_field.delete(0, END)
            brutto_field.delete(0, END)
            momsPercent_field.delete(0, END)
            day31.set('')
            day30.set('')
            day29.set('')
            sub1_drop.pack_forget()
            sub2_drop.pack_forget()
            v.set(0)

        def insert():
            # Take the data from the GUI and write to excel file
            if (receipt_name_field.get() == "" and
                    brutto_field.get() == "" and
                    momsPercent_field.get() == ""):
                print("empty input")

            else:
                current_row = wb.active.max_row + 1
                current_column = wb.active.max_column
                current_main_type = v.get()

                brutto = float(brutto_field.get())
                momsPercent = float(momsPercent_field.get())
                momsKr = float(brutto * (momsPercent / 100))
                netto = float(brutto - momsKr)
                verif_nr = current_row - 3
                subs1 = sub1.get()
                subs2 = sub2.get()

                wb.active.cell(row=current_row, column=2).value = receipt_name_field.get()

                if month == months_headers[1]:
                    wb.active.cell(row=current_row, column=1).value = day29.get()
                elif month == months_headers[0] or month == months_headers[2] or month == months_headers[4] or \
                        month == months_headers[6] or month == months_headers[7] or month == months_headers[9] or \
                        month == months_headers[11]:
                    wb.active.cell(row=current_row, column=1).value = day31.get()
                else:
                    wb.active.cell(row=current_row, column=1).value = day30.get()

                wb.active.cell(row=current_row, column=3).value = verif_nr



                # Inkomst
                if current_main_type == 1:
                    wb.active.cell(current_row, column=4).value = brutto
                    # Moms
                    wb.active.cell(current_row, column=6).value = momsKr
                    # Subtype
                    if subs1 == sub_type_headers1[1]:
                        # Column G = 7
                        wb.active.cell(current_row, column=7).value = netto
                    elif subs1 == sub_type_headers1[2]:
                        wb.active.cell(current_row, column=8).value = netto
                    elif subs1 == sub_type_headers1[3]:
                        wb.active.cell(current_row, column=9).value = netto
                    elif subs1 == sub_type_headers1[4]:
                        wb.active.cell(current_row, column=10).value = netto
                    elif subs1 == sub_type_headers1[5]:
                        wb.active.cell(current_row, column=11).value = netto

                # Utgift
                elif current_main_type == 2:
                    wb.active.cell(current_row, column=5).value = brutto
                    # Moms
                    if subs2 != sub_type_headers2[11]:
                        wb.active.cell(current_row, column=12).value = momsKr
                    else:
                        pass
                    # Subtype
                    if subs2 == sub_type_headers2[1]:
                        # Column M = 13
                        wb.active.cell(current_row, column=13).value = netto
                    elif subs2 == sub_type_headers2[2]:
                        wb.active.cell(current_row, column=14).value = netto
                    elif subs2 == sub_type_headers2[3]:
                        wb.active.cell(current_row, column=15).value = netto
                    elif subs2 == sub_type_headers2[4]:
                        wb.active.cell(current_row, column=16).value = netto
                    elif subs2 == sub_type_headers2[5]:
                        wb.active.cell(current_row, column=17).value = netto
                    elif subs2 == sub_type_headers2[6]:
                        wb.active.cell(current_row, column=18).value = netto
                    elif subs2 == sub_type_headers2[7]:
                        wb.active.cell(current_row, column=19).value = netto
                    elif subs2 == sub_type_headers2[8]:
                        wb.active.cell(current_row, column=20).value = netto
                    elif subs2 == sub_type_headers2[9]:
                        wb.active.cell(current_row, column=21).value = netto
                    elif subs2 == sub_type_headers2[10]:
                        wb.active.cell(current_row, column=22).value = netto
                    elif subs2 == sub_type_headers2[11]:
                        wb.active.cell(current_row, column=23).value = brutto
                    elif subs2 == sub_type_headers2[12]:
                        wb.active.cell(current_row, column=24).value = netto

            excel()
            wb.save(path)

            receipt_name_field.focus_set()

            clear()

        def ok():
            headvar = v.get()
            if headvar == 1:
                sub2_drop.pack_forget()
                sub1.set(sub_type_headers1[0])
                sub1_drop.pack(side=LEFT, pady=15, padx=5)
                sub1_drop.configure(width=20, font=11)
            elif headvar == 2:
                sub1_drop.pack_forget()
                sub2.set(sub_type_headers2[0])
                sub2_drop.pack(side=LEFT, pady=15, padx=5)
                sub2_drop.configure(width=40, font=11)

        sub1 = StringVar()
        sub1.set(sub_type_headers1[0])
        sub1_drop = OptionMenu(frame3_left_top, sub1, *sub_type_headers1)

        sub2 = StringVar()
        sub2.set(sub_type_headers2[0])
        sub2_drop = OptionMenu(frame3_left_top, sub2, *sub_type_headers2)

        # Set the config of GUI window
        root.geometry("500x550")

        # Create labels for every data entry
        Default_Label(mainframe, text=month).pack(padx=5, pady=5, side=TOP, anchor='n')
        v = IntVar()
        Radiobutton(mainframe, text="Inbetalning", font=14, variable=v, value=1, bg='light blue', command=ok).pack(
            side=LEFT,
            anchor='n',
            padx=70,
            pady=5)
        Radiobutton(mainframe, text="Utbetalning", font=14, variable=v, value=2, bg='light blue', command=ok).pack(
            side=RIGHT,
            anchor='n',
            padx=70,
            pady=5)

        Default_Label(frame2_left, text='Köpare, Säljare, Varumärke etc.:').pack(padx=5, pady=15, side=TOP)
        receipt_name_field = Entry(frame2_right, font=12)
        receipt_name_field.pack(padx=25, ipadx=15, pady=15, side=TOP)
        Default_Label(frame2_left, text='Dag:').pack(side=LEFT, pady=15, padx=5)

        day = StringVar()
        day31 = Combobox(frame2_left, width=3, textvariable=day, font=12)
        day30 = Combobox(frame2_left, width=3, textvariable=day, font=12)
        day29 = Combobox(frame2_left, width=3, textvariable=day, font=12)

        day31['values'] = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17',
                           '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31')
        day30['values'] = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17',
                           '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30')
        day29['values'] = ('1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17',
                           '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29')

        if month == months_headers[1]:
            day29.pack(side=LEFT, pady=15, padx=15)
        elif month == months_headers[0] or month == months_headers[2] or month == months_headers[4] or \
                month == months_headers[6] or month == months_headers[7] or month == months_headers[9] or \
                month == months_headers[11]:
            day31.pack(side=LEFT, pady=15, padx=15)
        else:
            day30.pack(side=LEFT, pady=15, padx=15)

        Default_Label(frame3_left_top, text="Sort:").pack(side=LEFT, pady=15, padx=5)

        Default_Label(frame3_left_bot, text="Pris:").pack(padx=5, pady=15, side=LEFT)
        brutto_field = Entry(frame3_left_bot, width=7, font=12)
        brutto_field.pack(padx=15, pady=15, side=LEFT)
        Default_Label(frame3_left_bot, text="Kr").pack(side=LEFT)

        Default_Label(frame4_top, text="Moms:").pack(side=LEFT, pady=15, padx=5)
        momsPercent_field = Entry(frame4_top, width=5, font=12)
        momsPercent_field.pack(side=LEFT, pady=15)
        Default_Label(frame4_top, text="%").pack(side=LEFT, pady=15, padx=5)

        # Save button
        Default_Button(frame4_bot, text="Spara", command=insert).pack(padx=50, pady=70, side=RIGHT, anchor='s')

        # Back Button
        Default_Button(frame4_bot, text="Tillbaka", command=months).pack(padx=50, pady=70, side=LEFT, anchor='s')

        # Bind method to call for the focus function
        receipt_name_field.bind("<Return>", focus1)
        brutto_field.bind("<Return>", focus2)
        excel()
        wb.save(path)

    months()

    root.mainloop()
